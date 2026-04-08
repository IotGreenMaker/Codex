#!/usr/bin/env python3
"""
Create starting Excel files for GBuddy application.
Generates g-buddy-data.xlsx with Plants, Watering_Log, Climate_Log, and Settings sheets.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import uuid
import json
import os

def generate_uuid():
    """Generate a UUID string."""
    return str(uuid.uuid4())

def create_header_style():
    """Create header cell style."""
    return {
        'font': Font(bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def apply_header_style(ws, row=1):
    """Apply header style to first row."""
    style = create_header_style()
    for cell in ws[row]:
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

def create_plants_sheet(wb):
    """Create Plants sheet with headers and sample data."""
    ws = wb.create_sheet('Plants')
    
    # Headers matching excel-storage.ts createPlantsSheet function
    headers = [
        'ID', 'Strain Name', 'Stage', 'Started At', 'Bloom Started At',
        'Light Schedule', 'Lights On', 'Lights Off', 'Light Type', 'Light Dimmer %',
        'Container Volume (L)', 'Media Volume (L)', 'Media Type', 'Grow Temp (°C)',
        'Grow Humidity (%)', 'Water Input (ml)', 'Water pH', 'Water EC',
        'Last Watered At', 'Watering Interval (days)', 'Outside Temp (°C)', 'Outside Humidity (%)'
    ]
    
    ws.append(headers)
    apply_header_style(ws)
    
    # Sample plant data
    sample_plants = [
        {
            'id': generate_uuid(),
            'strainName': 'Sour Diesel',
            'stage': 'Seedling',
            'startedAt': datetime.now().isoformat(),
            'bloomStartedAt': '',
            'lightSchedule': '18 / 6',
            'lightsOn': '06:00',
            'lightsOff': '00:00',
            'lightType': 'panel_100w',
            'lightDimmerPercent': 75,
            'containerVolumeL': 15,
            'mediaVolumeL': 13,
            'mediaType': 'Soil',
            'growTempC': 24,
            'growHumidity': 60,
            'waterInputMl': 500,
            'waterPh': 6.0,
            'waterEc': 1.0,
            'lastWateredAt': datetime.now().isoformat(),
            'wateringIntervalDays': 2,
            'outsideTempC': 20,
            'outsideHumidity': 50,
        },
        {
            'id': generate_uuid(),
            'strainName': 'Northern Lights',
            'stage': 'Veg',
            'startedAt': (datetime.now() - timedelta(days=14)).isoformat(),
            'bloomStartedAt': '',
            'lightSchedule': '18 / 6',
            'lightsOn': '06:00',
            'lightsOff': '00:00',
            'lightType': 'panel_100w',
            'lightDimmerPercent': 75,
            'containerVolumeL': 11,
            'mediaVolumeL': 9,
            'mediaType': 'Coco',
            'growTempC': 23,
            'growHumidity': 55,
            'waterInputMl': 400,
            'waterPh': 5.8,
            'waterEc': 1.2,
            'lastWateredAt': (datetime.now() - timedelta(days=1)).isoformat(),
            'wateringIntervalDays': 2,
            'outsideTempC': 18,
            'outsideHumidity': 45,
        },
    ]
    
    for plant in sample_plants:
        row = [
            plant['id'],
            plant['strainName'],
            plant['stage'],
            plant['startedAt'],
            plant['bloomStartedAt'],
            plant['lightSchedule'],
            plant['lightsOn'],
            plant['lightsOff'],
            plant['lightType'],
            plant['lightDimmerPercent'],
            plant['containerVolumeL'],
            plant['mediaVolumeL'],
            plant['mediaType'],
            plant['growTempC'],
            plant['growHumidity'],
            plant['waterInputMl'],
            plant['waterPh'],
            plant['waterEc'],
            plant['lastWateredAt'],
            plant['wateringIntervalDays'],
            plant['outsideTempC'],
            plant['outsideHumidity'],
        ]
        ws.append(row)
    
    # Set column widths
    column_widths = [36, 20, 10, 24, 24, 14, 10, 10, 14, 14, 18, 16, 12, 14, 16, 14, 10, 10, 24, 20, 16, 16]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return ws

def create_watering_sheet(wb):
    """Create Watering_Log sheet with headers."""
    ws = wb.create_sheet('Watering_Log')
    
    headers = ['ID', 'Plant ID', 'Timestamp', 'Amount (ml)', 'pH', 'EC', 'Runoff pH', 'Runoff EC']
    ws.append(headers)
    apply_header_style(ws)
    
    # Set column widths
    column_widths = [36, 36, 24, 12, 8, 8, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return ws

def create_climate_sheet(wb):
    """Create Climate_Log sheet with headers."""
    ws = wb.create_sheet('Climate_Log')
    
    headers = ['ID', 'Plant ID', 'Timestamp', 'Temp (°C)', 'Humidity (%)']
    ws.append(headers)
    apply_header_style(ws)
    
    # Set column widths
    column_widths = [36, 36, 24, 12, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return ws

def create_settings_sheet(wb):
    """Create Settings sheet with default settings."""
    ws = wb.create_sheet('Settings')
    
    headers = ['Key', 'Value']
    ws.append(headers)
    apply_header_style(ws)
    
    # Default settings
    settings = [
        {'key': 'activePlantId', 'value': ''},
        {'key': 'locale', 'value': 'en'},
    ]
    
    for setting in settings:
        ws.append([setting['key'], setting['value']])
    
    # Set column widths
    column_widths = [20, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return ws

def main():
    """Main function to create Excel files."""
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(output_dir, 'g-buddy-data.xlsx')
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create sheets in order
    create_plants_sheet(wb)
    create_watering_sheet(wb)
    create_climate_sheet(wb)
    create_settings_sheet(wb)
    
    # Save workbook
    wb.save(output_file)
    
    print(f'✓ Created Excel file: {output_file}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Plants: 2 sample entries')
    print(f'  Watering_Log: headers only')
    print(f'  Climate_Log: headers only')
    print(f'  Settings: default configuration')
    
    # Also try to load from existing plants-state.json if it exists
    json_file = os.path.join(output_dir, 'plants-state.json')
    if os.path.exists(json_file):
        print(f'\nℹ Found existing plants-state.json')
        print(f'  To migrate data, run: python migrate-json-to-excel.py')

if __name__ == '__main__':
    main()